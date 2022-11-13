from linkedin_scraper.objects import Experience #type: ignore
import openpyxl, time #type: ignore
import pandas as pd #type: ignore
from datetime import datetime #type: ignore
from linkedin_scraper import Person, Company, actions #type: ignore
from selenium import webdriver #type: ignore
from selenium.webdriver.common.by import By 
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import os, re, sys
from column_cell import column_value
from blank_check import blank_cell
from fuzzywuzzy import fuzz
from progress.bar import Bar


def linkedIn_sales_datas(person_url, browser):
    name, company_name, job_title, experience_scrap_det = '', '', '', {}
    browser.get(person_url)
    try:
        WebDriverWait(browser, 45).until(EC.presence_of_all_elements_located((By.XPATH,'//span[@class="artdeco-button__text"]')))
    except Exception as e:
        print('Error :', e)
    
    for _ in range(3):
        try:
            body = browser.find_element_by_css_selector('body')
            body.send_keys(Keys.PAGE_DOWN)
            time.sleep(1)
        except Exception as e:
            print('Error =>', e)

    soup = BeautifulSoup(browser.page_source, 'lxml')
    profile_details = soup.find('div', {'class': 'profile-topcard-person-entity__content min-width inline-block'})
    experience_details = soup.find('div', {'id': 'profile-experience'}) 

    if profile_details is not None:
        p_name = profile_details.find('span', {'class': 'profile-topcard-person-entity__name t-24 t-black t-bold'})
        p_job = profile_details.find('dd', {'class': 'mt2'})
        
        if p_name is not None:
            name = p_name.text.strip()
        if p_job is not None:
            job_title = p_job.text.strip()
            backup_jt = job_title

    if experience_details is not None:
        exp_lists = soup.find_all('li', {'class': 'profile-position display-flex align-items-flex-start'})

        for ind, exp in enumerate(exp_lists, 1):
            comp_name, emp_date, duration = '', '', ''
            experience_lists = [ i.strip() for i in exp.text.strip().split('\n') if i.strip() != '' ]

            if 'Company Name' in experience_lists:
                comp_ind = experience_lists.index('Company Name')
                comp_name = experience_lists[comp_ind+1]

                try:
                    if ind == 1:
                        tit_index = experience_lists.index('Title')
                        job_title = experience_lists[tit_index+1]
                except Exception as e:
                    try:
                        if ind == 1:
                            job_title = experience_lists[comp_ind-1]
                    except:
                        job_title = backup_jt

            if 'Dates Employed' in experience_lists:
                date_ind = experience_lists.index('Dates Employed')
                emp_date = experience_lists[date_ind+1]

            if 'Employment Duration' in experience_lists:
                dur_ind = experience_lists.index('Employment Duration')
                duration = experience_lists[dur_ind+1]

            experience_scrap_det[ind] = [comp_name,emp_date,duration]

    if experience_scrap_det:
        company_name = experience_scrap_det[1][0]

    # company_name, job_title, experience_scrap_det
    result = {
        'Title URL': person_url,
        'Name': name,
        'Company Name': company_name,
        'Job Title': job_title,
        'Experience': experience_scrap_det,
    }
    
    return result

def linkedIn_person_data(person_url, browser):
    name, company_name, job_title, experience_scrap_det = '', '', '', {}
    browser.get(person_url)
    try:
        WebDriverWait(browser, 45).until(EC.presence_of_all_elements_located((By.XPATH,'//span[@class="artdeco-button__text"]')))
    except Exception as e:
        print('Error :', e)
    
    for _ in range(3):
        try:
            body = browser.find_element_by_css_selector('body')
            body.send_keys(Keys.PAGE_DOWN)
            time.sleep(1)
        except Exception as e:
            print('Error =>', e)

    soup = BeautifulSoup(browser.page_source, 'lxml')
    profile_details = soup.find('div', {'class': 'mt2 relative'})
    experience_details = soup.find('section', {'id': 'experience-section'})
    
    with open('exp.txt', 'w') as ff:
        ff.write(str(experience_details))

    if profile_details is not None:
        p_name = profile_details.find('h1', {'class': 'text-heading-xlarge inline t-24 v-align-middle break-words'})
        p_job = profile_details.find('div', {'class': 'text-body-medium break-words'})
        p_comp = profile_details.find('div', {'aria-label': 'Current company'})
        if p_name is not None:
            name = p_name.text.strip()
        if p_job is not None:
            job_title = p_job.text.strip()
            backup_jt = job_title

    if experience_details is not None:
        exp_lists = experience_details.find_all('li')
        for ind, exp in enumerate(exp_lists, 1):
            comp_name, emp_date, duration = '', '', ''
            experience_lists = [ i.strip() for i in exp.text.strip().split('\n') if i.strip() != '' ]
            # print(experience_lists)
            if 'Company Name' in experience_lists:
                comp_ind = experience_lists.index('Company Name')
                comp_name = experience_lists[comp_ind+1]
                try:
                    if ind == 1:
                        tit_index = experience_lists.index('Title')
                        job_title = experience_lists[tit_index+1]
                except:
                    try:
                        if ind == 1:
                            job_title = experience_lists[comp_ind-1]
                    except:
                        job_title = backup_jt 
            if 'Dates Employed' in experience_lists:
                date_ind = experience_lists.index('Dates Employed')
                emp_date = experience_lists[date_ind+1]

            if 'Employment Duration' in experience_lists:
                dur_ind = experience_lists.index('Employment Duration')
                duration = experience_lists[dur_ind+1]
            
            experience_scrap_det[ind] = [comp_name, emp_date, duration]
    
    if experience_scrap_det:
        company_name = experience_scrap_det[1][0]

    # company_name, job_title, experience_scrap_det
    result = {
        'Title URL': person_url,
        'Name': name,
        'Company Name': company_name,
        'Job Title': job_title,
        'Experience': experience_scrap_det,
    }

    return result

def linkedIn_login(email_id, password, browser):
    try:
        url = "https://www.linkedin.com/login"
        browser.get(url)
        WebDriverWait(browser, 45).until(EC.element_to_be_clickable((By.XPATH,'//input[@id="username"]'))).send_keys(email_id)
        WebDriverWait(browser, 45).until(EC.element_to_be_clickable((By.XPATH,'//input[@id="password"]'))).send_keys(password)
        WebDriverWait(browser, 45).until(EC.element_to_be_clickable((By.XPATH,'//button[@type="submit"]'))).click()
    except Exception as e:
        print(f'RAPBot Failed Login Issue - {e}')
    
    chk_curr_url = str(browser.current_url).lower()

    if 'login-submit' in chk_curr_url or 'checkpoint' in chk_curr_url:
        print('RAPBot Failed Login Issue..')
        browser.close()
        return 0

def linkedIn_company_data(company_url, browser):
    # Company(linkedin_url=None, name=None, about_us=None, website=None, headquarters=None, 
    # founded=None, company_type=None, company_size=None, specialties=None, showcase_pages=[], 
    # affiliated_companies=[], driver=None, scrape=True, get_employees=True)
    
    company = Company(linkedin_url=company_url, driver=browser, 
                        close_on_complete=False, get_employees=False)
    company_datas = {   
                        "Company URL": company_url,
                        "Company Name": company.name,
                        "Industry": company.industry,
                        "Company Size": company.company_size,
                    }

    return company_datas
    
def read_excel_lead_scraped_datas(excel_path):

    outputs, results = [], []

    for excel_file in os.listdir(excel_path):
        if 'lead' in excel_file and '.xlsx' in excel_file:
            full_path = os.path.join(excel_path, excel_file)
            data_dict = pd.read_excel(full_path).to_dict()
            result = {
						"Title URL": [],
						"Name": [],
						"Company Name": [],
						"Job Title": [],
						"Experience": [],
					}

            for d_dict in data_dict:
                for dd in data_dict[d_dict]:
                    if 'title url' == d_dict.lower():
                        result['Title URL'].append(data_dict[d_dict][dd])
                    if 'name' == d_dict.lower():
                        result['Name'].append(data_dict[d_dict][dd])
                    if 'job title' == d_dict.lower():
                        result['Job Title'].append(data_dict[d_dict][dd])
                    if 'company name' == d_dict.lower():
                        result['Company Name'].append(data_dict[d_dict][dd])
                    if 'experience' == d_dict.lower():
                        experience = {}
                        for ind, dat in enumerate(eval(data_dict[d_dict][dd]), 1):
                            experience[ind] = dat
                        result['Experience'].append(experience)
                if result not in outputs:
                    outputs.append(result)

    if outputs:
        for output in outputs:
            data_size = len(output['Title URL'])
            for i in range(data_size):
                data = {
						"Title URL": '',
						"Name": '',
						"Company Name": '',
						"Job Title": '',
						"Experience": {},
					}
                data['Title URL'] = output['Title URL'][i]
                data['Name'] = output['Name'][i]
                data['Company Name'] = output['Company Name'][i]
                data['Job Title'] = output['Job Title'][i]
                data['Experience'] = output['Experience'][i]
                results.append(data)

    return results

def check_lead_data_in_excel(chk_url, scrap_file):
    lead_datas = f"{scrap_file}\\scrapping_data"
    try:
        if os.path.exists(lead_datas):
            for data in read_excel_lead_scraped_datas(lead_datas):
                if data['Title URL'] == chk_url:
                    return data
        return None
    except Exception as e:
        print(f'RAPBot Failed - {e}')
        return None
        
def read_excel_comp_scraped_datas(excel_path):

	results = []

	for excel_file in os.listdir(excel_path):
		if 'company' in excel_file and '.xlsx' in excel_file:
			full_path = os.path.join(excel_path, excel_file)
			data_dict = pd.read_excel(full_path).to_dict()
			if data_dict:
				data_size = len(data_dict['Company URL'])
				for i in range(data_size):
					data = {
						"Company URL": '',
						"Company Name": '',
						"Industry": '',
						"Company Size": '',
					}
					data['Company URL'] = data_dict['Company URL'][i]
					data['Company Name'] = data_dict['Company Name'][i]
					data['Industry'] = data_dict['Industry'][i]
					data['Company Size'] = data_dict['Company Size'][i]
					results.append(data)
	
	return results

def check_comp_data_in_excel(chk_url, scrap_file):
    lead_datas = f"{scrap_file}\\scrapping_data"
    try:
        if os.path.exists(lead_datas):
            for data in read_excel_comp_scraped_datas(lead_datas):
                if data['Company URL'] == chk_url:
                    return data
        return None
    except Exception as e:
        print(f'RAPBot Failed - {e}')
        return None

def linkedin_logout(browser):
    # WebDriverWait(browser, 45).until(EC.element_to_be_clickable((By.XPATH, "//button[@id='ember32']"))).click()
    logout_url = "https://www.linkedin.com/m/logout"
    browser.get(logout_url)

def lead_comp(main_file, email_id, pwd, chrome_path, scrap_file):
    for file_na in os.listdir(main_file):
        if "main qc file" in file_na.lower():
            main_excel = openpyxl.load_workbook(f"{main_file}\\{file_na}")
            main_sheet = main_excel.worksheets[0]
            qualified = column_value(f"{main_file}\\{file_na}", "qualified")
            primary = column_value(f"{main_file}\\{file_na}", "primary reason")
            jtl = column_value(f"{main_file}\\{file_na}", "job title link")
            full = column_value(f"{main_file}\\{file_na}", "full name")
            comments = column_value(f"{main_file}\\{file_na}", "bot comments")
            ten_link = column_value(f"{main_file}\\{file_na}", "tenure on linkedin")
            jt = column_value(f"{main_file}\\{file_na}", "title")
            comp_name = column_value(f"{main_file}\\{file_na}", "company name")
            csl = column_value(f"{main_file}\\{file_na}", "company size link")
            inds = column_value(f"{main_file}\\{file_na}", "industry")
            emplo_size = column_value(f"{main_file}\\{file_na}", "employee size")
            blank = blank_cell(f"{main_file}\\{file_na}")
            print(blank)
            if blank != {}:
                for k, v in blank.items():
                    quali = main_sheet.cell(row = k, column = comments).value
                    if quali != None:
                        if quali.lower() == "no":
                            pass
                    else:
                        # main_sheet.cell(row = k, column = qualified).value = "No"
                        main_sheet.cell(row = k, column = comments).value = f"Blank Field : {v}"
        
            browser = webdriver.Chrome(chrome_path)
            browser.maximize_window()
            try:
                # email_id, password = "@gmail.com", "@260783"
                chk_login = linkedIn_login(email_id, pwd, browser)
                if chk_login == 0:
                    return 'Login Issue [Username or Password] incorrect'
            except Exception as e:
                print(f'RAPBot Login Failed reason - {e}')
            
            lead_linkedin_datas, comp_linkedin_datas = [], []
            b = Bar('Loading : ', max=main_sheet.max_row)

            try:
                for row in range(2, main_sheet.max_row+1):
                    b.next()
                    print()
                    job_title_link = str(main_sheet.cell(row=row, column=jtl).value).strip()
                    company_link = str(main_sheet.cell(row=row, column=csl).value).strip()
                    print(job_title_link, company_link)
                    chk_excel_flag = False
                    if 'none' == job_title_link.lower() or '-' == job_title_link:
                        main_sheet.cell(row=row, column=comments).value = 'Blank Job: Title Link'
                        print('Blank URL..')
                    else:
                        if 'sales/people' in job_title_link.lower():
                            try:
                                exc_data = check_lead_data_in_excel(job_title_link, scrap_file)
                                if exc_data is not None:
                                    profile_details = exc_data
                                    chk_excel_flag = True
                                else:
                                    profile_details = linkedIn_sales_datas(str(job_title_link), browser)
                            except Exception as e:
                                print(f'RAPBot Failed reason - {e}')
                                profile_details = {}
                        else:
                            try:
                                exc_data = check_lead_data_in_excel(job_title_link, scrap_file)
                                if exc_data is not None:
                                    profile_details = exc_data
                                    chk_excel_flag = True
                                else:
                                    profile_details = linkedIn_person_data(str(job_title_link), browser)
                            except Exception as e:
                                print(f'RAPBot Failed reason - {e}')
                                profile_details = {}
                     
                        if not profile_details:
                            bot_command = main_sheet.cell(row=row, column=comments)
                            bot_command.value = str(str(bot_command.value) + '/Bot Failed to Fetch LinkedIn Profile').strip('None')
                        else:
                            print(profile_details)
                            if profile_details['Name'] == '' and profile_details['Company Name'] == '' and profile_details['Job Title'] == '' and profile_details['Experience'] == {}:
                                qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                    main_sheet.cell(row=row, column=qualified).value = 'No'
                                    main_sheet.cell(row=row, column=primary).value = 'Incorrect Lead URL'
                                print('Bot Failed to Fetch LinkedIn Profile - May be Page Not Found')
                            else:
                                try:
                                    if profile_details.get('Name') is not None:
                                        fulname = main_sheet.cell(row=row, column=full).value
                                        qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                        primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                        prof_name = re.sub('[^a-zA-Z]', '', str(profile_details.get('Name')).lower().strip())
                                        fulname = re.sub('[^a-zA-Z]', '', str(fulname).lower().strip())

                                        name_per = fuzz.partial_ratio(fulname, prof_name)

                                        if name_per >= 70:
                                            print(fulname, prof_name, name_per, 'matched')
                                        elif name_per >= 60 and name_per < 70:
                                            print(fulname, prof_name, name_per, 'partial matched')
                                            bot_command = main_sheet.cell(row=row, column=comments)
                                            bot_command.value = str(str(bot_command.value) + '/Partial Full Name').strip('None')
                                        else:
                                            print(fulname, prof_name, 'unmatched')
                                            if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                main_sheet.cell(row=row, column=qualified).value = 'No'
                                                main_sheet.cell(row=row, column=primary).value = 'Invalid full name'
                                except Exception as e:
                                    print(f'RAPBot Failed reason - {e}')
                                
                                try:
                                    if profile_details.get('Job Title') is not None:
                                        job_tit = str(main_sheet.cell(row=row, column=jt).value).lower().strip().replace('&', 'and')
                                        prof_job_tit = str(profile_details.get('Job Title')).lower().strip().replace('&', 'and')
                                        qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                        primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                        
                                        job_per = fuzz.partial_ratio(job_tit, prof_job_tit)

                                        if job_per >= 75:
                                            print(prof_job_tit, f'{job_per} matched')
                                        elif job_per >= 70 and job_per < 75:
                                            print(prof_job_tit, f'{job_per} partial matched')
                                            bot_command = main_sheet.cell(row=row, column=comments)
                                            bot_command.value = str(str(bot_command.value) + '/Partial Job Title').strip('None')
                                        else:
                                            print(prof_job_tit, f'{job_per} unmatched')
                                            if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                main_sheet.cell(row=row, column=qualified).value = 'No'
                                                main_sheet.cell(row=row, column=primary).value = 'Invalid Job Title'
                                except Exception as e:
                                    print(f'RAPBot Failed reason - {e}')

                                try:    
                                    if profile_details.get('Company Name') is not None:
                                        comp = str(main_sheet.cell(row=row, column=comp_name).value).lower().strip()
                                        prof_comp = str(profile_details.get('Company Name')).lower().strip()

                                        if ' at ' in comp:
                                            comp = comp.split(' at ')[0].strip()
                                        elif '@' in comp:
                                            comp = comp.split('@')[0].strip()

                                        if ' at ' in prof_comp:
                                            prof_comp = prof_comp.split(' at ')[0].strip()
                                        elif '@' in prof_comp:
                                            prof_comp = prof_comp.split('@')[0].strip()

                                        qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                        primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                        
                                        comp_per = fuzz.partial_ratio(comp, prof_comp)
                                        if comp_per >= 75:
                                            print(prof_comp, f'{comp_per} matched')
                                        elif comp_per >= 70 and comp_per < 75:
                                            print(prof_comp, f'{comp_per} partial matched')
                                            bot_command = main_sheet.cell(row=row, column=comments)
                                            bot_command.value = str(str(bot_command.value) + '/Partial Company Name').strip('None')
                                        else:
                                            print(prof_comp, f'{comp_per} unmatched')
                                            if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                main_sheet.cell(row=row, column=qualified).value = 'No'
                                                main_sheet.cell(row=row, column=primary).value = 'Invalid Company Name'
                                except Exception as e:
                                    print(f'RAPBot Failed reason  - {e}')
                                
                                try:
                                    if profile_details.get('Experience'):
                                        tenure = main_sheet.cell(row=row, column=ten_link).value
                                        tenure = str(tenure).lower().strip().replace(' ', '')

                                        if len(profile_details.get('Experience')) > 1:
                                            pres_cnt, present_list, c_present_list = 0, [], []
                                            experiences = profile_details.get('Experience')
                                            # print(profile_details.get('Experience'))
                                            for experience in experiences:
                                                if experiences[experience][1:] not in c_present_list:
                                                    present_list.append(experiences[experience])
                                                    c_present_list.append(experiences[experience][1:])
                                            
                                            s_comp_name, cnt_p = str(present_list[0][0]).lower(), 0
                                            for i in present_list:
                                                if str(i[0]).strip():
                                                    if 'present' in str(i[1]).lower():
                                                        if s_comp_name in str(i[0]).lower() and cnt_p == 0:
                                                            pres_cnt += 1
                                                            cnt_p += 1
                                                        if s_comp_name in str(i[0]).lower() and cnt_p > 0:
                                                            continue
                                                        if s_comp_name.lower().strip() not in str(i[0]).lower():
                                                            pres_cnt += 1
                                                        
                                            if pres_cnt > 1:
                                                print('Present :', pres_cnt)
                                                qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                                primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                                if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                    main_sheet.cell(row=row, column=qualified).value = 'No'
                                                    main_sheet.cell(row=row, column=primary).value = 'Dual Employement'
                                        
                                        p_tenure1 = profile_details.get('Experience')[1][1].replace(' ', '')
                                        p_tenure2 = profile_details.get('Experience')[1][2].replace(' ', '')
                                        ext_yrs = re.findall('[0-9]+\s?yr', p_tenure2)
                                        if ext_yrs:
                                            yrs = re.findall('[0-9]+', ext_yrs[0])[0]
                                            yrs = int(re.sub('[^0-9]', '', yrs))
                                            if yrs >= 7:
                                                print('YEAR :', yrs)
                                                qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                                primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                                if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                    main_sheet.cell(row=row, column=qualified).value = 'No'
                                                    main_sheet.cell(row=row, column=primary).value = 'Invalid Experience'
                                        profile_details['Experience'] = present_list
                                except Exception as e:
                                    print(f'RAPBot Failed reason - {e}')
                        if profile_details:
                            if chk_excel_flag == False:
                                lead_linkedin_datas.append(profile_details)

                    print('-'*70)
                    # company details scraping
                    chk_excel_flag = False
                    if 'none' == company_link.lower() or '-' == company_link:
                        main_sheet.cell(row=row, column=comments).value = 'Blank Company Link'
                        print('Blank URL..')
                    else:                
                        if 'sales/company' in company_link.lower():
                            org_url = company_link.split('company')
                            org_url = "https://www.linkedin.com/company/" + org_url[-1].strip('/').split('/')[0]
                            company_link = org_url

                        try:
                            comp_link = company_link
                            if comp_link.endswith('/about'): 
                                comp_link = comp_link.strip('/about')
                            elif comp_link.endswith('/about/'):
                                comp_link = comp_link.strip('/about/')

                            exc_data = check_comp_data_in_excel(comp_link, scrap_file)
                            if exc_data is not None:
                                company_details = exc_data
                                chk_excel_flag = True
                            elif comp_linkedin_datas != []:
                                for comp_lin_data in comp_linkedin_datas:
                                    try:
                                        if comp_lin_data['Company URL'] == comp_link:
                                            company_details = comp_lin_data
                                            break
                                    except Exception as e:
                                        print(e)
                                        company_details = linkedIn_company_data(comp_link, browser)
                                else:
                                    company_details = linkedIn_company_data(comp_link, browser)        
                            else:
                                company_details = linkedIn_company_data(comp_link, browser)
                            print('Company Details :', company_details)
                        except Exception as e:
                            print(f'RAPBot Failed reason - {e}')
                            company_details = {}
                        
                        if not company_details:
                            bot_command = main_sheet.cell(row=row, column=comments)
                            bot_command.value = str(str(bot_command.value) + '/Bot Failed to Fetch LinkedIn Company').strip('None')
                        else:            
                            if company_details['Industry'] == '' and company_details['Company Size'] == '':
                                qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                    main_sheet.cell(row=row, column=qualified).value = 'No'
                                    main_sheet.cell(row=row, column=primary).value = 'Incorrect Company URL'
                                print('Bot Failed to Fetch LinkedIn Company - May be Page Not Found')
                            else:
                                try:    
                                    if company_details.get('Company Name') is not None:
                                        comp = str(main_sheet.cell(row=row, column=comp_name).value).lower().strip()
                                        c_comp = str(company_details.get('Company Name')).lower().strip()
                                        
                                        if ' at ' in comp:
                                            comp = comp.split(' at ')[0].strip()
                                        elif '@' in comp:
                                            comp = comp.split('@')[0].strip()

                                        if ' at ' in c_comp:
                                            c_comp = c_comp.split(' at ')[0].strip()
                                        elif '@' in c_comp:
                                            c_comp = c_comp.split('@')[0].strip()

                                        qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                        primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()

                                        comp_per = fuzz.partial_ratio(comp, c_comp)
                                        if comp_per >= 65:
                                            print(c_comp, f'{comp_per} matched')
                                        elif comp_per >= 60 and comp_per < 65:
                                            print(c_comp, f'{comp_per} partial matched')
                                            bot_command = main_sheet.cell(row=row, column=comments)
                                            bot_command.value = str(str(bot_command.value) + '/Partial Company Name').strip('None')
                                        else:
                                            print(c_comp, f'{comp_per} unmatched')
                                            if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                main_sheet.cell(row=row, column=qualified).value = 'No'
                                                main_sheet.cell(row=row, column=primary).value = 'Invalid Company Name'
                                except Exception as e:
                                    print(f'RAPBot Failed reason  - {e}')

                                try:
                                    if company_details.get('Industry') is not None:
                                        industry = str(main_sheet.cell(row=row, column=inds).value).lower().strip()
                                        comp_indus = str(company_details.get('Industry')).lower().strip()

                                        ind_per = fuzz.partial_ratio(industry, comp_indus)
                                        if ind_per >= 75:
                                            print(comp_indus, f'{ind_per} matched')
                                        elif ind_per >= 70 and ind_per < 75:
                                            print(comp_indus, f'{ind_per} partial matched')
                                            bot_command = main_sheet.cell(row=row, column=comments)
                                            bot_command.value = str(str(bot_command.value) + '/Partial Industry').strip('None')
                                        else:
                                            print(industry, f'{ind_per} unmatched')
                                            qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                            primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                            if qualified_field in ['none', ''] and primary_field in ['none', '']:
                                                main_sheet.cell(row=row, column=qualified).value = 'No'
                                                main_sheet.cell(row=row, column=primary).value = 'Invalid Industry'
                                except Exception as e:
                                    print(f'RAPBot Failed reason - {e}')

                                try:
                                    if company_details.get('Company Size') is not None:
                                        emp_size = str(main_sheet.cell(row=row, column=emplo_size).value).strip().replace(',', '').replace(' ', '')
                                        comp_size = str(company_details.get('Company Size')).lower().strip().replace(',', '').replace(' ', '')
                                        if comp_size in emp_size:
                                            print(emp_size, comp_size, 'matched')
                                        elif emp_size in comp_size:
                                            print(emp_size, comp_size, 'matched')
                                        else:
                                            print(emp_size, comp_size, 'unmatched')
                                            qualified_field = str(main_sheet.cell(row=row, column=qualified).value).lower().strip()
                                            primary_field = str(main_sheet.cell(row=row, column=primary).value).lower().strip()
                                            bot_cmt = str(main_sheet.cell(row=row, column=comments).value).strip().lower()
                                            if bot_cmt in ['none', ''] and bot_cmt in ['none', '']:
                                                # main_sheet.cell(row=row, column=qualified).value = 'No'
                                                # main_sheet.cell(row=row, column=primary).value = 'Invalid Company Size'
                                                main_sheet.cell(row=row, column=comments).value = 'Invalid Company Size'
                                except Exception as e:
                                    print(f'RAPBot Failed reason - {e}')

                        if company_details:
                            if chk_excel_flag == False:
                                # print(company_details)
                                comp_linkedin_datas.append(company_details)
                    print('-'*70)

                try:
                    linkedin_logout(browser)
                except: pass
                
                browser.close()
                main_excel.save(f"{main_file}\\{file_na}")
            except Exception as e:
                print(f'RAPBot Failed reason - {e}')

            try:
                os.makedirs(f"{scrap_file}\\scrapping_data", exist_ok=True)
                todays_date = datetime.today().strftime('%d-%m-%Y %M%S')
                if lead_linkedin_datas:
                    df1 = pd.DataFrame(lead_linkedin_datas)
                    df1.to_excel(f'{scrap_file}\\scrapping_data\\lead_scraped_datas_{todays_date}.xlsx', index=False)
            except Exception as e:
                print(f'RAPBot Failed reason - {e}')
            
            try:
                todays_date = datetime.today().strftime('%d-%m-%Y %M%S')
                if comp_linkedin_datas:
                    df2 = pd.DataFrame(comp_linkedin_datas)
                    df2.to_excel(f'{scrap_file}\\scrapping_data\\company_scraped_datas_{todays_date}.xlsx', index=False)
            except Exception as e:
                print(f'RAPBot Failed reason - {e}')

            try:
                b.finish()
            except: pass
            
# main_file = "data\\Main QC"
# email = ''
# pwd = ''
# chromedriver = "C:\\Program Files (x86)\\Google\\chrome-driver\\96\\chromedriver.exe"

# lead_comp(main_file, email, pwd, chromedriver, "data\\26112")